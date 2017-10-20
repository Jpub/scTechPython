# openpyxl에서 필요한 함수(load_workbook)를 import
from openpyxl import load_workbook
import numpy as np  # NumPy도 import

# WorkBook을 연다
wb = load_workbook(filename='Sample1.xlsx', read_only=True,
                   data_only=True, use_iterators=True)
# WorkSheet를 이름으로 지정
ws = wb['온도변화']

# 데이터를 저장할 NumPy ndarray를 미리 만들어둔다
Nrow = 11
time_vec = np.zeros(Nrow)
temp_vec = np.zeros(Nrow)

# 데이터 읽기
for i, row in enumerate(ws.iter_rows(row_offset=1)):
    time_vec[i] = row[0].value
    temp_vec[i] = row[1].value
